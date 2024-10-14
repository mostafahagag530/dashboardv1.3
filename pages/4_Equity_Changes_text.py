import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
import os
def process_excel(file_path, sheet_name=0, cell_reference='B53'):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")

    # Load the workbook
    workbook = load_workbook(file_path, data_only=True)
    
    # Determine the sheet to load
    if isinstance(sheet_name, int):
        if sheet_name >= len(workbook.sheetnames) or sheet_name < 0:
            raise ValueError(f"Sheet index {sheet_name} is out of range.")
        sheet = workbook.worksheets[sheet_name]
    else:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")
        sheet = workbook[sheet_name]

    # Get value from a specific cell
    cell_value = sheet[cell_reference].value if cell_reference else None
    return cell_value

# Function to preprocess the data
def preprocess_data_df(excel_file_path):
    # Load the Excel file to access specific cell ranges
    xls = pd.ExcelFile(excel_file_path)

    # Define the sheet name
    sheet_name = xls.sheet_names[0]  # Assuming you want the first sheet
    B53_Text = process_excel(excel_file_path)
    # Read data from ranges with no header
    df_M = pd.read_excel(excel_file_path, sheet_name=sheet_name, usecols="M", skiprows=4, nrows=9, header=None)
    df_P = pd.read_excel(excel_file_path, sheet_name=sheet_name, usecols="P", skiprows=4, nrows=9, header=None)

    Equity_df = pd.concat([df_M, df_P], axis=1)

    # Reset index if needed
    Equity_df.reset_index(drop=True, inplace=True)
    
    # Extract the first and last date from the first and last rows
    Equity_df.columns = ["Date", "value"]
    Equity_df["Date"].iloc[0] = pd.to_datetime(Equity_df["Date"].iloc[0]).strftime('%d %m %y')
    Equity_df["Date"].iloc[-1] = pd.to_datetime(Equity_df["Date"].iloc[-1]).strftime('%d %m %y')
    start_date = pd.to_datetime(Equity_df["Date"].iloc[0]).strftime('%d %m')
    end_date = pd.to_datetime(Equity_df["Date"].iloc[-1]).strftime('%m')

    # Renaming columns with the extracted date range
    Equity_df["Date"] = Equity_df["Date"].apply(lambda x: f"{x} {start_date} : {end_date}")
    Equity_df["B53_Text"] = f"{start_date} : {end_date} \n {B53_Text}"
    
    return Equity_df

st.set_page_config(layout="wide")

@st.cache_data
def load_and_preprocess(filepath):
    processed_df = preprocess_data_df(filepath)
    return processed_df

# Streamlit app title
st.title(" Waterfall Chart of Equity Changes and His B53 text ")

# Sidebar for file uploader and date range selection
st.sidebar.header("Upload and Filter Data")

uploaded_files = st.sidebar.file_uploader("Choose Excel files", type="xlsx", accept_multiple_files=True)

# Initialize the DataFrame variable
df = pd.DataFrame()

if uploaded_files:
    all_dfs = []
    for uploaded_file in uploaded_files:
        # Save the uploaded file temporarily
        with open(os.path.join("./", uploaded_file.name), "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        # Process the saved file
        df_temp = load_and_preprocess(os.path.join("./", uploaded_file.name))
        all_dfs.append(df_temp)
    
    # Concatenate all DataFrames into one
    df = pd.concat(all_dfs, ignore_index=True)

    # Select the columns to display
    selected_columns = st.sidebar.multiselect(
        "Select columns to display", 
        df.columns.tolist(), 
        default=df.columns.tolist()  # Set default to all columns
    )
    
    # Ensure at least one column is selected
    if selected_columns:
        num_rows_to_display = st.sidebar.slider(
            "Select number of rows to display", 
            min_value=1, 
            max_value=len(df), 
            value=10
        )

        # Display the filtered dataframe
        st.dataframe(df[selected_columns].head(num_rows_to_display), use_container_width=True)
    else:
        st.warning("Please select at least one column to display.")

    Equity_df=df
    Equity_df= Equity_df.drop_duplicates().reset_index(drop=True)
    # # Show the figure
    def create_theme():
        return {
            'title_font': dict(size=24, family='Arial, sans-serif', color="black"),
            'axis_font': dict(size=12, family='Arial, sans-serif', color="black"),
            'colors': ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728'],  # Default colors
            'bg_color': '#f5f5f5',
            'grid_color': 'rgba(0, 0, 0, 0.1)',
            'paper_bgcolor': 'white',
        }

    theme = create_theme()

    # Ensure the first column of the DataFrame is treated as strings (for categorical x-axis)
    Equity_df[Equity_df.columns[0]] = Equity_df[Equity_df.columns[0]].astype(str)

    # Define colors for the waterfall bars (green for positive, red for negative) from the theme
    colors = ['#2ca02c' if value >= 0 else '#d62728' for value in Equity_df[Equity_df.columns[1]]]

    # Create the waterfall chart
    fig = go.Figure(go.Waterfall(
        x=Equity_df[Equity_df.columns[0]],  # Categorical text for the x-axis
        y=Equity_df[Equity_df.columns[1]],  # Numerical value for the y-axis
        measure=["total"] + ["relative"] * (len(Equity_df[Equity_df.columns[1]]) - 2) + ["total"],  # First and last as total
        textposition="outside",
        text=[f'{value:,.2f}' for value in Equity_df[Equity_df.columns[1]]],  # Format numbers with commas
        hovertext=[f'Value: {value:,.2f}' for value in Equity_df[Equity_df.columns[1]]],  # Optional hover text
        # marker=dict(color=colors)  # Use defined colors for the bars
    ))

    # Update layout with theme settings and improved design
    fig.update_layout(
        title='Waterfall Chart of Equity Changes',
        title_x=0.3,  # Center the title
        title_font=theme['title_font'],  # Use theme for title font
        xaxis_title='Categories',
        yaxis_title='Portfolio Value',
        
        # X-axis styling based on theme
        xaxis=dict(
            title_font=theme['axis_font'],  # Use theme for axis font
            type='category',  # Ensure x-axis is treated as categorical
            tickfont=theme['axis_font'],  # Font for x-axis ticks
            gridcolor=theme['grid_color'],  # Grid color from theme
        ),
        
        # Y-axis styling based on theme
        yaxis=dict(
            title_font=theme['axis_font'],  # Use theme for axis font
            tickfont=theme['axis_font'],  # Font for y-axis ticks
            gridcolor=theme['grid_color'],  # Grid color from theme
            zeroline=False  # Hide zero line
        ),
        
        # Set background and paper color from the theme
        plot_bgcolor=theme['bg_color'],  # Set plot background color
        paper_bgcolor=theme['paper_bgcolor'],  # Set paper background color
        
        # Global font settings for all text
        font=theme['axis_font'],  # Font for all text
        
        # Update margins for better spacing
        margin=dict(l=40, r=40, t=80, b=40),
        width=800,  # Set the width of the figure
        height=600, 
    
    )    # Define figure size


    # Drop NaN values and reset the index again
    # Equity_df["B53_Text"] = Equity_df["B53_Text"].dropna().reset_index(drop=True)
    # st.table(Equity_df["B53_Text"])


                # Display the chart with full width inside the loop
    with st.container():
                    st.plotly_chart(fig, use_container_width=True)